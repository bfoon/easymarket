from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Avg
from django.core.paginator import Paginator
from django.forms import modelformset_factory
from logistics.models import Shipment
from logistics.forms import ShipmentForm
from orders.models import Order, ShippingAddress
from django.utils import timezone
from datetime import timedelta
from django.db import transaction
from django.contrib.auth import get_user_model
import csv
import json
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.utils import get_column_letter

from .models import (
    Shipment, ShipmentBox, BoxItem, Driver, Vehicle,
    Warehouse, LogisticOffice
)
from orders.models import Order, OrderItem
from .forms import ShipmentBoxForm, BoxItemForm

User = get_user_model()

class ShipmentListView(LoginRequiredMixin, ListView):
    model = Shipment
    template_name = 'logistics/shipment_list.html'
    context_object_name = 'shipments'
    paginate_by = 20

    def get_queryset(self):
        queryset = Shipment.objects.select_related(
            'shipping_address', 'warehouse', 'driver', 'vehicle', 'order'
        ).order_by('-created_at')

        # Search functionality
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(id__icontains=search_query) |
                Q(shipping_address__address__icontains=search_query) |
                Q(driver__user__first_name__icontains=search_query) |
                Q(driver__user__last_name__icontains=search_query) |
                Q(order__id__icontains=search_query)  # Added order ID search
            )

        # Status filter
        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Order status filter
        order_status_filter = self.request.GET.get('order_status')
        if order_status_filter:
            queryset = queryset.filter(order__status=order_status_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = Shipment.STATUS_CHOICES
        context['order_status_choices'] = [
            ('pending', 'Pending'),
            ('processing', 'Processing'),
            ('shipped', 'Shipped'),
            ('delivered', 'Delivered'),
            ('cancelled', 'Cancelled'),
        ]
        context['current_status'] = self.request.GET.get('status', '')
        context['current_order_status'] = self.request.GET.get('order_status', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class ShipmentDetailView(LoginRequiredMixin, DetailView):
    model = Shipment
    template_name = 'logistics/shipment_detail.html'
    context_object_name = 'shipment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        shipment = self.get_object()
        context['boxes'] = shipment.boxes.prefetch_related('items__order_item__product')
        context['total_boxes'] = shipment.boxes.count()

        # Add order status information
        if shipment.order:
            order = shipment.order
            context['order'] = order
            context['order_status'] = order.status
            context['order_status_display'] = order.get_status_display()
            context['is_order_delivered'] = order.status == 'delivered'
            context['can_mark_delivered'] = (
                    order.status in ['shipped'] and
                    shipment.status == 'shipped'
            )
            context['delivered_date'] = order.delivered_date
            # Make boxes read-only if order is delivered
            context['boxes_readonly'] = order.status == 'delivered'
        else:
            context['order'] = None
            context['boxes_readonly'] = False

        return context


class ShipmentCreateView(LoginRequiredMixin, CreateView):
    model = Shipment
    template_name = 'logistics/shipment_form.html'
    form_class = ShipmentForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['processing_orders_count'] = Order.objects.filter(status='processing').count()
        return context

    def get_success_url(self):
        return reverse_lazy('logistics:shipment_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # You can also assign request.user or other meta info here
        response = super().form_valid(form)
        messages.success(self.request, 'Shipment created successfully!')
        return response

    def form_invalid(self, form):
        # Log all field errors in the console for debugging
        print("\n=== Shipment Form Errors ===")
        for field, errors in form.errors.items():
            print(f"{field}: {errors}")

        # Optional: Set a user-facing message
        messages.error(self.request, "There was an error creating the shipment. Please review the form.")
        return super().form_invalid(form)

def ajax_addresses_for_order(request):
    order_id = request.GET.get('order_id')
    addresses = ShippingAddress.objects.filter(order_id=order_id)
    data = {
        'addresses': [
            {'id': a.id, 'display': str(a)} for a in addresses
        ]
    }
    return JsonResponse(data)

def ajax_orders_for_address(request):
    address_id = request.GET.get('address_id')
    address = ShippingAddress.objects.filter(id=address_id).first()
    orders = Order.objects.filter(shippingaddress=address) if address else []
    data = {
        'orders': [
            {'id': o.id, 'display': f'#{o.id} - {o.created_at.strftime("%Y-%m-%d")}'}
            for o in orders
        ]
    }
    return JsonResponse(data)

class ShipmentUpdateView(LoginRequiredMixin, UpdateView):
    model = Shipment
    template_name = 'logistics/shipment_form.html'
    fields = [
        'warehouse', 'driver', 'vehicle', 'logistic_office',
        'collect_time', 'estimated_dropoff_time', 'weight_kg',
        'size_cubic_meters', 'material_type', 'shipment_type', 'packing_type',
        'container_type', 'status'
    ]

    def get_success_url(self):
        return reverse_lazy('logistics:shipment_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Shipment updated successfully!')
        return super().form_valid(form)


class DriverListView(LoginRequiredMixin, ListView):
    model = Driver
    template_name = 'logistics/driver_list.html'
    context_object_name = 'drivers'
    paginate_by = 20

    def get_queryset(self):
        queryset = Driver.objects.select_related('user').prefetch_related(
            'vehicle_set', 'shipment_set'
        ).order_by('user__first_name')

        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(license_number__icontains=search_query) |
                Q(phone__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add driver statistics
        for driver in context['drivers']:
            driver.active_shipments_count = driver.shipment_set.exclude(status='shipped').count()
            driver.total_shipments_count = driver.shipment_set.count()
            driver.vehicle_count = driver.vehicle_set.count()

        return context


class DriverDetailView(LoginRequiredMixin, DetailView):
    model = Driver
    template_name = 'logistics/driver_detail.html'
    context_object_name = 'driver'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        driver = self.get_object()

        # Get vehicles with computed status
        vehicles = driver.vehicle_set.all()
        for vehicle in vehicles:
            vehicle.active_shipments_count = vehicle.shipment_set.exclude(status='shipped').count()

        context['vehicles'] = vehicles
        context['recent_shipments'] = Shipment.objects.filter(
            driver=driver
        ).select_related('shipping_address', 'vehicle').order_by('-created_at')[:10]

        # Add computed statistics
        context['total_shipments'] = driver.shipment_set.count()
        context['active_shipments'] = driver.shipment_set.exclude(status='shipped').count()
        context['completed_shipments'] = driver.shipment_set.filter(status='shipped').count()
        context['vehicle_count'] = driver.vehicle_set.count()

        return context


class DriverCreateView(LoginRequiredMixin, CreateView):
    model = Driver
    form_class = None  # We'll set this in get_form_class
    template_name = 'logistics/driver_form.html'

    def get_form_class(self):
        from .forms import DriverForm
        return DriverForm

    def get_success_url(self):
        return reverse_lazy('logistics:driver_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Driver created successfully!')
        return super().form_valid(form)


class DriverUpdateView(LoginRequiredMixin, UpdateView):
    model = Driver
    form_class = None
    template_name = 'logistics/driver_form.html'

    def get_form_class(self):
        from .forms import DriverForm
        return DriverForm

    def get_success_url(self):
        return reverse_lazy('logistics:driver_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Driver {form.instance.user.get_full_name()} updated successfully!')
        return super().form_valid(form)


class VehicleCreateView(LoginRequiredMixin, CreateView):
    model = Vehicle
    form_class = None
    template_name = 'logistics/vehicle_form.html'

    def get_form_class(self):
        from .forms import VehicleForm
        return VehicleForm

    def get_success_url(self):
        return reverse_lazy('logistics:vehicle_list')

    def form_valid(self, form):
        messages.success(self.request, f'Vehicle {form.instance.plate_number} created successfully!')
        return super().form_valid(form)


class VehicleUpdateView(LoginRequiredMixin, UpdateView):
    model = Vehicle
    form_class = None
    template_name = 'logistics/vehicle_form.html'

    def get_form_class(self):
        from .forms import VehicleForm
        return VehicleForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object:
            context['active_shipments_count'] = self.object.shipment_set.exclude(status='shipped').count()
        return context

    def get_success_url(self):
        return reverse_lazy('logistics:vehicle_list')

    def form_valid(self, form):
        messages.success(self.request, f'Vehicle {form.instance.plate_number} updated successfully!')
        return super().form_valid(form)


class VehicleDetailView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'logistics/vehicle_detail.html'
    context_object_name = 'vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehicle = self.get_object()

        # Get shipment statistics
        context['total_shipments'] = vehicle.shipment_set.count()
        context['active_shipments'] = vehicle.shipment_set.exclude(status='shipped').count()
        context['completed_shipments'] = vehicle.shipment_set.filter(status='shipped').count()

        # Get recent shipments
        context['recent_shipments'] = vehicle.shipment_set.select_related(
            'shipping_address', 'driver'
        ).order_by('-created_at')[:10]

        return context


class WarehouseCreateView(LoginRequiredMixin, CreateView):
    model = Warehouse
    form_class = None
    template_name = 'logistics/warehouse_form.html'

    def get_form_class(self):
        from .forms import WarehouseForm
        return WarehouseForm

    def get_success_url(self):
        return reverse_lazy('logistics:warehouse_list')

    def form_valid(self, form):
        messages.success(self.request, f'Warehouse {form.instance.name} created successfully!')
        return super().form_valid(form)


class WarehouseUpdateView(LoginRequiredMixin, UpdateView):
    model = Warehouse
    form_class = None
    template_name = 'logistics/warehouse_form.html'

    def get_form_class(self):
        from .forms import WarehouseForm
        return WarehouseForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object:
            context['active_shipments_count'] = self.object.shipment_set.exclude(status='shipped').count()
            context['completed_shipments_count'] = self.object.shipment_set.filter(status='shipped').count()
        return context

    def get_success_url(self):
        return reverse_lazy('logistics:warehouse_list')

    def form_valid(self, form):
        messages.success(self.request, f'Warehouse {form.instance.name} updated successfully!')
        return super().form_valid(form)


class WarehouseDetailView(LoginRequiredMixin, DetailView):
    model = Warehouse
    template_name = 'logistics/warehouse_detail.html'
    context_object_name = 'warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse = self.get_object()

        # Get shipment statistics
        context['total_shipments'] = warehouse.shipment_set.count()
        context['active_shipments'] = warehouse.shipment_set.exclude(status='shipped').count()
        context['completed_shipments'] = warehouse.shipment_set.filter(status='shipped').count()

        # Get recent shipments
        context['recent_shipments'] = warehouse.shipment_set.select_related(
            'shipping_address', 'driver', 'vehicle'
        ).order_by('-created_at')[:10]

        return context


class LogisticOfficeCreateView(LoginRequiredMixin, CreateView):
    model = LogisticOffice
    form_class = None
    template_name = 'logistics/logistic_office_form.html'

    def get_form_class(self):
        from .forms import LogisticOfficeForm
        return LogisticOfficeForm

    def get_success_url(self):
        return reverse_lazy('logistics:logistic_office_list')

    def form_valid(self, form):
        messages.success(self.request, f'Logistic Office {form.instance.name} created successfully!')
        return super().form_valid(form)


class LogisticOfficeUpdateView(LoginRequiredMixin, UpdateView):
    model = LogisticOffice
    form_class = None
    template_name = 'logistics/logistic_office_form.html'

    def get_form_class(self):
        from .forms import LogisticOfficeForm
        return LogisticOfficeForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object:
            context['active_shipments_count'] = self.object.shipment_set.exclude(status='shipped').count()
            context['completed_shipments_count'] = self.object.shipment_set.filter(status='shipped').count()
        return context

    def get_success_url(self):
        return reverse_lazy('logistics:logistic_office_list')

    def form_valid(self, form):
        messages.success(self.request, f'Logistic Office {form.instance.name} updated successfully!')
        return super().form_valid(form)


class LogisticOfficeListView(LoginRequiredMixin, ListView):
    model = LogisticOffice
    template_name = 'logistics/logistic_office_list.html'
    context_object_name = 'offices'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add office statistics
        total_shipments_sum = 0
        active_shipments_sum = 0
        completed_shipments_sum = 0

        for office in context['offices']:
            office.total_shipments_count = office.shipment_set.count()
            office.active_shipments_count = office.shipment_set.exclude(status='shipped').count()
            office.completed_shipments_count = office.shipment_set.filter(status='shipped').count()

            total_shipments_sum += office.total_shipments_count
            active_shipments_sum += office.active_shipments_count
            completed_shipments_sum += office.completed_shipments_count

        context['total_shipments_sum'] = total_shipments_sum
        context['active_shipments_sum'] = active_shipments_sum
        context['completed_shipments_sum'] = completed_shipments_sum

        return context


class VehicleListView(LoginRequiredMixin, ListView):
    model = Vehicle
    template_name = 'logistics/vehicle_list.html'
    context_object_name = 'vehicles'
    paginate_by = 20

    def get_queryset(self):
        queryset = Vehicle.objects.select_related('driver__user').prefetch_related(
            'shipment_set'
        ).order_by('plate_number')

        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(plate_number__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(driver__user__first_name__icontains=search_query) |
                Q(driver__user__last_name__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add vehicle statistics
        active_count = 0
        available_count = 0

        for vehicle in context['vehicles']:
            vehicle.active_shipments_count = vehicle.shipment_set.exclude(status='shipped').count()
            vehicle.total_shipments_count = vehicle.shipment_set.count()
            # Calculate current load (you might want to implement this based on your business logic)
            vehicle.current_load = 0  # Placeholder - implement based on active shipments
            vehicle.current_usage_percent = 0  # Placeholder

            if vehicle.active_shipments_count > 0:
                active_count += 1
            else:
                available_count += 1

        context['active_vehicles_count'] = active_count
        context['available_vehicles_count'] = available_count

        return context


class WarehouseListView(LoginRequiredMixin, ListView):
    model = Warehouse
    template_name = 'logistics/warehouse_list.html'
    context_object_name = 'warehouses'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add warehouse statistics
        total_shipments_sum = 0
        active_shipments_sum = 0
        completed_shipments_sum = 0

        for warehouse in context['warehouses']:
            warehouse.total_shipments_count = warehouse.shipment_set.count()
            warehouse.active_shipments_count = warehouse.shipment_set.exclude(status='shipped').count()
            warehouse.completed_shipments_count = warehouse.shipment_set.filter(status='shipped').count()

            total_shipments_sum += warehouse.total_shipments_count
            active_shipments_sum += warehouse.active_shipments_count
            completed_shipments_sum += warehouse.completed_shipments_count

        context['total_shipments_sum'] = total_shipments_sum
        context['active_shipments_sum'] = active_shipments_sum
        context['completed_shipments_sum'] = completed_shipments_sum

        return context


def update_shipment_status(request, pk):
    """AJAX view to update shipment status"""
    if request.method == 'POST':
        shipment = get_object_or_404(Shipment, pk=pk)
        new_status = request.POST.get('status')

        if new_status in dict(Shipment.STATUS_CHOICES):
            shipment.status = new_status
            shipment.save()

            if new_status == 'shipped':
                shipment.mark_as_shipped()

            return JsonResponse({
                'success': True,
                'message': f'Shipment status updated to {shipment.get_status_display()}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request'})


def generate_box_label(request, shipment_id, box_id):
    """Generate QR code label for a box"""
    shipment = get_object_or_404(Shipment, pk=shipment_id)
    box = get_object_or_404(ShipmentBox, pk=box_id, shipment=shipment)

    box.generate_qr_label()
    box.save()

    messages.success(request, f'QR label generated for Box #{box.box_number}')
    return redirect('logistics:shipment_detail', pk=shipment_id)


def mark_order_as_delivered(request, shipment_pk):
    """Mark the associated order as delivered and update shipment status"""
    if request.method == 'POST':
        shipment = get_object_or_404(Shipment, pk=shipment_pk)

        if not shipment.order:
            return JsonResponse({
                'success': False,
                'message': 'No order associated with this shipment'
            })

        # Check if shipment is shipped
        if shipment.status != 'shipped':
            return JsonResponse({
                'success': False,
                'message': 'Shipment must be shipped before marking as delivered'
            })

        # Update order status to delivered
        order = shipment.order
        old_status = order.status
        order.status = 'delivered'

        # Set delivered date if not already set
        if not order.delivered_date:
            order.delivered_date = timezone.now()

        order.save()

        # Update shipment status (you might want to add a 'delivered' status to Shipment model)
        # For now, we'll keep it as 'shipped' since the shipment itself is complete

        # Create order status history
        from orders.models import OrderStatusHistory
        OrderStatusHistory.objects.create(
            order=order,
            status='delivered',
            changed_by=request.user,
            notes=f'Marked as delivered via logistics app - Shipment #{shipment.id}'
        )

        return JsonResponse({
            'success': True,
            'message': f'Order #{order.id} marked as delivered successfully',
            'new_status': order.get_status_display(),
            'delivered_date': order.delivered_date.strftime('%B %d, %Y at %I:%M %p') if order.delivered_date else None
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def get_shipment_order_status(request, shipment_pk):
    """Get the current order status for a shipment"""
    shipment = get_object_or_404(Shipment, pk=shipment_pk)

    if not shipment.order:
        return JsonResponse({
            'has_order': False,
            'message': 'No order associated with this shipment'
        })

    order = shipment.order
    return JsonResponse({
        'has_order': True,
        'order_id': order.id,
        'order_status': order.status,
        'order_status_display': order.get_status_display(),
        'can_mark_delivered': order.status == 'shipped' and shipment.status == 'shipped',
        'delivered_date': order.delivered_date.strftime('%B %d, %Y at %I:%M %p') if order.delivered_date else None,
        'is_delivered': order.status == 'delivered'
    })


class DashboardView(LoginRequiredMixin, ListView):
    template_name = 'logistics/dashboard.html'

    def get(self, request, *args, **kwargs):
        context = {
            'total_shipments': Shipment.objects.count(),
            'pending_shipments': Shipment.objects.filter(status='pending').count(),
            'in_transit_shipments': Shipment.objects.filter(status='in_transit').count(),
            'shipped_shipments': Shipment.objects.filter(status='shipped').count(),
            'total_drivers': Driver.objects.count(),
            'total_vehicles': Vehicle.objects.count(),
            'total_warehouses': Warehouse.objects.count(),
            'recent_shipments': Shipment.objects.select_related(
                'shipping_address', 'driver', 'vehicle'
            ).order_by('-created_at')[:10],
        }
        return render(request, self.template_name, context)


# Box Management Views
class ShipmentBoxCreateView(LoginRequiredMixin, CreateView):
    model = ShipmentBox
    form_class = ShipmentBoxForm
    template_name = 'logistics/box_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.shipment = get_object_or_404(Shipment, pk=kwargs['shipment_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['shipment'] = self.shipment
        return kwargs

    def form_valid(self, form):
        form.instance.shipment = self.shipment
        messages.success(self.request, f'Box #{form.instance.box_number} created successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.shipment.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['shipment'] = self.shipment
        context['action'] = 'Create'
        return context


class ShipmentBoxUpdateView(LoginRequiredMixin, UpdateView):
    model = ShipmentBox
    form_class = ShipmentBoxForm
    template_name = 'logistics/box_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['shipment'] = self.object.shipment
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f'Box #{form.instance.box_number} updated successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.object.shipment.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['shipment'] = self.object.shipment
        context['action'] = 'Edit'
        return context


class ShipmentBoxDeleteView(LoginRequiredMixin, DeleteView):
    model = ShipmentBox
    template_name = 'logistics/box_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        shipment_pk = self.object.shipment.pk
        box_number = self.object.box_number

        success_url = reverse('logistics:shipment_detail', kwargs={'pk': shipment_pk})
        result = super().delete(request, *args, **kwargs)

        messages.success(request, f'Box #{box_number} deleted successfully!')
        return result

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.object.shipment.pk})


class BoxItemCreateView(LoginRequiredMixin, CreateView):
    model = BoxItem
    form_class = BoxItemForm
    template_name = 'logistics/box_item_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.box = get_object_or_404(ShipmentBox, pk=kwargs['box_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['box'] = self.box
        return kwargs

    def form_valid(self, form):
        form.instance.box = self.box
        messages.success(self.request, 'Item added to box successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.box.shipment.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['box'] = self.box
        context['shipment'] = self.box.shipment
        context['action'] = 'Add'
        return context


class BoxItemUpdateView(LoginRequiredMixin, UpdateView):
    model = BoxItem
    form_class = BoxItemForm
    template_name = 'logistics/box_item_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['box'] = self.object.box
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Box item updated successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.object.box.shipment.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['box'] = self.object.box
        context['shipment'] = self.object.box.shipment
        context['action'] = 'Edit'
        return context


class BoxItemDeleteView(LoginRequiredMixin, DeleteView):
    model = BoxItem
    template_name = 'logistics/box_item_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        shipment_pk = self.object.box.shipment.pk

        success_url = reverse('logistics:shipment_detail', kwargs={'pk': shipment_pk})
        result = super().delete(request, *args, **kwargs)

        messages.success(request, 'Item removed from box successfully!')
        return result

    def get_success_url(self):
        return reverse('logistics:shipment_detail', kwargs={'pk': self.object.box.shipment.pk})


def manage_shipment_boxes(request, shipment_pk):
    """Bulk manage boxes for a shipment"""
    shipment = get_object_or_404(Shipment, pk=shipment_pk)

    BoxFormSet = modelformset_factory(
        ShipmentBox,
        form=ShipmentBoxForm,
        extra=3,  # Allow 3 new boxes by default
        can_delete=True
    )

    if request.method == 'POST':
        formset = BoxFormSet(request.POST, queryset=shipment.boxes.all())

        if formset.is_valid():
            instances = formset.save(commit=False)

            for instance in instances:
                instance.shipment = shipment
                instance.save()

            # Delete marked instances
            for instance in formset.deleted_objects:
                instance.delete()

            messages.success(request, 'Boxes updated successfully!')
            return redirect('logistics:shipment_detail', pk=shipment.pk)
    else:
        formset = BoxFormSet(queryset=shipment.boxes.all())

    # Modify each form to include the shipment reference
    for form in formset:
        if hasattr(form, 'instance') and form.instance.pk is None:
            form.instance.shipment = shipment

    context = {
        'shipment': shipment,
        'formset': formset,
    }
    return render(request, 'logistics/manage_boxes.html', context)


def get_or_create_unassigned_driver():
    """Get or create a special 'Unassigned' driver for vehicles without active drivers"""
    try:
        # Try to find existing unassigned driver
        unassigned_user = User.objects.filter(
            username='unassigned_driver',
            is_active=False
        ).first()

        if not unassigned_user:
            # Create unassigned user
            unassigned_user = User.objects.create_user(
                username='unassigned_driver',
                email='unassigned@system.local',
                first_name='Unassigned',
                last_name='Driver',
                is_active=False  # Mark as inactive so it doesn't appear in normal listings
            )

        # Get or create the driver
        unassigned_driver, created = Driver.objects.get_or_create(
            user=unassigned_user,
            defaults={
                'phone': 'N/A',
                'license_number': 'UNASSIGNED'
            }
        )

        return unassigned_driver
    except Exception as e:
        print(f"Error creating unassigned driver: {e}")
        return None


class AssignmentDashboardView(LoginRequiredMixin, ListView):
    """Dashboard for managing driver-vehicle assignments"""
    template_name = 'logistics/assignment_dashboard.html'

    def get(self, request, *args, **kwargs):
        # Get the unassigned driver
        unassigned_driver = get_or_create_unassigned_driver()

        # Get vehicles with their assignment status (excluding unassigned driver)
        vehicles = Vehicle.objects.select_related('driver__user').annotate(
            active_shipments_count=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
        ).order_by('plate_number')

        # Get drivers with their vehicle counts (excluding unassigned driver)
        drivers = Driver.objects.select_related('user').filter(
            user__is_active=True
        ).annotate(
            vehicle_count=Count('vehicle', filter=~Q(vehicle__driver=unassigned_driver)),
            active_shipments_count=Count('vehicle__shipment',
                                         filter=Q(vehicle__shipment__status__in=['pending', 'in_transit']))
        ).order_by('user__first_name')

        # Get unassigned vehicles (assigned to unassigned driver)
        unassigned_vehicles = vehicles.filter(driver=unassigned_driver) if unassigned_driver else []

        # Get drivers without vehicles (excluding unassigned driver)
        drivers_without_vehicles = drivers.filter(vehicle_count=0)

        # Get real assigned vehicles (not assigned to unassigned driver)
        assigned_vehicles_count = vehicles.exclude(
            driver=unassigned_driver).count() if unassigned_driver else vehicles.count()

        context = {
            'vehicles': vehicles.exclude(driver=unassigned_driver) if unassigned_driver else vehicles,
            'drivers': drivers,
            'unassigned_vehicles': unassigned_vehicles,
            'drivers_without_vehicles': drivers_without_vehicles,
            'total_vehicles': vehicles.count(),
            'assigned_vehicles': assigned_vehicles_count,
            'active_drivers': drivers.filter(vehicle_count__gt=0).count(),
            'total_drivers': drivers.count(),
        }
        return render(request, self.template_name, context)


class VehicleAssignmentListView(LoginRequiredMixin, ListView):
    """List all vehicles and their current driver assignments"""
    model = Vehicle
    template_name = 'logistics/vehicle_assignment_list.html'
    context_object_name = 'vehicles'
    paginate_by = 20

    def get_queryset(self):
        unassigned_driver = get_or_create_unassigned_driver()

        queryset = Vehicle.objects.select_related('driver__user').annotate(
            active_shipments_count=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
        ).order_by('plate_number')

        # Apply filters
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(plate_number__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(driver__user__first_name__icontains=search_query) |
                Q(driver__user__last_name__icontains=search_query)
            )

        assignment_status = self.request.GET.get('assignment_status')
        if assignment_status == 'assigned':
            queryset = queryset.exclude(driver=unassigned_driver) if unassigned_driver else queryset.filter(
                driver__isnull=False)
        elif assignment_status == 'unassigned':
            queryset = queryset.filter(driver=unassigned_driver) if unassigned_driver else queryset.filter(
                driver__isnull=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['assignment_status'] = self.request.GET.get('assignment_status', '')
        context['drivers'] = Driver.objects.select_related('user').filter(user__is_active=True)
        context['unassigned_driver'] = get_or_create_unassigned_driver()
        return context


def assign_vehicle_to_driver(request):
    """AJAX view to assign a vehicle to a driver"""
    if request.method == 'POST':
        vehicle_id = request.POST.get('vehicle_id')
        driver_id = request.POST.get('driver_id')

        try:
            vehicle = Vehicle.objects.get(id=vehicle_id)

            # Handle unassignment (when driver_id is empty)
            if not driver_id:
                unassigned_driver = get_or_create_unassigned_driver()
                if not unassigned_driver:
                    return JsonResponse({
                        'success': False,
                        'message': 'Unable to create unassigned driver placeholder'
                    })
                driver = unassigned_driver
            else:
                driver = Driver.objects.get(id=driver_id)

            # Check if vehicle is currently in use for active shipments
            active_shipments = vehicle.shipment_set.filter(status__in=['pending', 'in_transit']).count()
            if active_shipments > 0 and vehicle.driver != driver:
                return JsonResponse({
                    'success': False,
                    'message': f'Vehicle {vehicle.plate_number} has {active_shipments} active shipments and cannot be reassigned.'
                })

            old_driver = vehicle.driver
            vehicle.driver = driver
            vehicle.save()

            # Create appropriate message
            if driver.user.username == 'unassigned_driver':
                message = f'Vehicle {vehicle.plate_number} unassigned from driver'
                driver_name = 'Unassigned'
            else:
                message = f'Vehicle {vehicle.plate_number} assigned to {driver.user.get_full_name()}'
                driver_name = driver.user.get_full_name()

            if old_driver and old_driver != driver and old_driver.user.username != 'unassigned_driver':
                message += f' (previously assigned to {old_driver.user.get_full_name()})'

            return JsonResponse({
                'success': True,
                'message': message,
                'vehicle_id': vehicle.id,
                'driver_name': driver_name
            })

        except (Vehicle.DoesNotExist, Driver.DoesNotExist) as e:
            return JsonResponse({
                'success': False,
                'message': 'Vehicle or driver not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error assigning vehicle: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def unassign_vehicle(request, vehicle_id):
    """Unassign a vehicle from its current driver"""
    if request.method == 'POST':
        try:
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)

            # Check for active shipments
            active_shipments = vehicle.shipment_set.filter(status__in=['pending', 'in_transit']).count()
            if active_shipments > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'Vehicle {vehicle.plate_number} has {active_shipments} active shipments and cannot be unassigned.'
                })

            # Get unassigned driver
            unassigned_driver = get_or_create_unassigned_driver()
            if not unassigned_driver:
                return JsonResponse({
                    'success': False,
                    'message': 'Unable to create unassigned driver placeholder'
                })

            old_driver = vehicle.driver
            vehicle.driver = unassigned_driver
            vehicle.save()

            message = f'Vehicle {vehicle.plate_number} unassigned'
            if old_driver and old_driver.user.username != 'unassigned_driver':
                message += f' from {old_driver.user.get_full_name()}'

            return JsonResponse({
                'success': True,
                'message': message
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error unassigning vehicle: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def quick_assign_vehicle(request):
    """Quick assignment form for immediate vehicle assignment"""
    if request.method == 'POST':
        vehicle_id = request.POST.get('vehicle_id')
        driver_id = request.POST.get('driver_id')

        try:
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
            driver = get_object_or_404(Driver, id=driver_id)

            # Check for active shipments
            active_shipments = vehicle.shipment_set.filter(status__in=['pending', 'in_transit']).count()
            if active_shipments > 0 and vehicle.driver != driver:
                messages.error(request,
                               f'Vehicle {vehicle.plate_number} has {active_shipments} active shipments and cannot be reassigned.')
                return redirect('logistics:assignment_dashboard')

            vehicle.driver = driver
            vehicle.save()

            messages.success(request,
                             f'Vehicle {vehicle.plate_number} successfully assigned to {driver.user.get_full_name()}')

        except Exception as e:
            messages.error(request, f'Error assigning vehicle: {str(e)}')

    return redirect('logistics:assignment_dashboard')


def get_unassigned_vehicles_ajax(request):
    """Get all unassigned vehicles via AJAX"""
    try:
        unassigned_driver = get_or_create_unassigned_driver()

        if unassigned_driver:
            vehicles = Vehicle.objects.filter(driver=unassigned_driver)
        else:
            vehicles = Vehicle.objects.none()

        vehicles_data = []
        for vehicle in vehicles:
            vehicles_data.append({
                'id': vehicle.id,
                'plate_number': vehicle.plate_number,
                'model': vehicle.model,
                'capacity_kg': vehicle.capacity_kg
            })

        return JsonResponse({
            'success': True,
            'vehicles': vehicles_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving unassigned vehicles: {str(e)}'
        })


def bulk_assign_vehicles(request):
    """Bulk assignment of multiple vehicles"""
    if request.method == 'POST':
        vehicle_ids = request.POST.getlist('vehicle_ids')
        driver_ids = request.POST.getlist('driver_ids')

        if len(vehicle_ids) != len(driver_ids):
            messages.error(request, 'Number of vehicles must match number of drivers')
            return redirect('logistics:assignment_dashboard')

        try:
            unassigned_driver = get_or_create_unassigned_driver()

            with transaction.atomic():
                assignments_made = 0
                for vehicle_id, driver_id in zip(vehicle_ids, driver_ids):
                    vehicle = Vehicle.objects.get(id=vehicle_id)

                    # Handle unassignment
                    if not driver_id:
                        if not unassigned_driver:
                            continue
                        driver = unassigned_driver
                    else:
                        driver = Driver.objects.get(id=driver_id)

                    # Check for active shipments
                    active_shipments = vehicle.shipment_set.filter(status__in=['pending', 'in_transit']).count()
                    if active_shipments > 0 and vehicle.driver != driver:
                        messages.warning(request,
                                         f'Skipped vehicle {vehicle.plate_number} - has active shipments')
                        continue

                    vehicle.driver = driver
                    vehicle.save()
                    assignments_made += 1

                if assignments_made > 0:
                    messages.success(request, f'Successfully assigned {assignments_made} vehicles')
                else:
                    messages.warning(request, 'No vehicles were assigned')

        except Exception as e:
            messages.error(request, f'Error in bulk assignment: {str(e)}')

    return redirect('logistics:assignment_dashboard')


# Update the existing DriverDetailView in your views.py
class DriverDetailView(LoginRequiredMixin, DetailView):
    model = Driver
    template_name = 'logistics/driver_detail.html'
    context_object_name = 'driver'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        driver = self.get_object()
        unassigned_driver = get_or_create_unassigned_driver()

        # Get vehicles with computed status, excluding unassigned driver
        vehicles = driver.vehicle_set.all()
        assigned_vehicles = vehicles.exclude(driver=unassigned_driver) if unassigned_driver else vehicles

        for vehicle in assigned_vehicles:
            vehicle.active_shipments_count = vehicle.shipment_set.exclude(status='shipped').count()

        context['vehicles'] = assigned_vehicles
        context['assigned_vehicles_count'] = assigned_vehicles.count()
        context['recent_shipments'] = Shipment.objects.filter(
            driver=driver
        ).select_related('shipping_address', 'vehicle').order_by('-created_at')[:10]

        # Add computed statistics
        context['total_shipments'] = driver.shipment_set.count()
        context['active_shipments'] = driver.shipment_set.exclude(status='shipped').count()
        context['completed_shipments'] = driver.shipment_set.filter(status='shipped').count()
        context['vehicle_count'] = assigned_vehicles.count()

        return context


class DriverAssignmentListView(LoginRequiredMixin, ListView):
    """List all drivers and their assigned vehicles"""
    model = Driver
    template_name = 'logistics/driver_assignment_list.html'
    context_object_name = 'drivers'
    paginate_by = 20

    def get_queryset(self):
        unassigned_driver = get_or_create_unassigned_driver()

        queryset = Driver.objects.select_related('user').filter(
            user__is_active=True
        ).annotate(
            vehicle_count=Count('vehicle', filter=~Q(vehicle__driver=unassigned_driver) if unassigned_driver else Q()),
            active_shipments_count=Count('vehicle__shipment',
                                         filter=Q(vehicle__shipment__status__in=['pending', 'in_transit']))
        ).order_by('user__first_name')

        # Apply filters
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(license_number__icontains=search_query) |
                Q(phone__icontains=search_query)
            )

        vehicle_status = self.request.GET.get('vehicle_status')
        if vehicle_status == 'with_vehicles':
            queryset = queryset.filter(vehicle_count__gt=0)
        elif vehicle_status == 'without_vehicles':
            queryset = queryset.filter(vehicle_count=0)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['vehicle_status'] = self.request.GET.get('vehicle_status', '')

        unassigned_driver = get_or_create_unassigned_driver()
        context['unassigned_vehicles'] = Vehicle.objects.filter(driver=unassigned_driver) if unassigned_driver else []
        return context


def get_driver_vehicles_ajax(request, driver_id):
    """Get vehicles assigned to a specific driver via AJAX"""
    try:
        driver = get_object_or_404(Driver, id=driver_id)
        unassigned_driver = get_or_create_unassigned_driver()

        # Get vehicles assigned to this driver (excluding unassigned driver)
        vehicles = driver.vehicle_set.all()
        if unassigned_driver:
            vehicles = vehicles.exclude(driver=unassigned_driver)

        vehicles_data = []
        for vehicle in vehicles:
            active_shipments = vehicle.shipment_set.filter(status__in=['pending', 'in_transit']).count()
            vehicles_data.append({
                'id': vehicle.id,
                'plate_number': vehicle.plate_number,
                'model': vehicle.model,
                'capacity_kg': vehicle.capacity_kg,
                'active_shipments': active_shipments,
                'status': 'In Use' if active_shipments > 0 else 'Available'
            })

        return JsonResponse({
            'success': True,
            'driver_name': driver.user.get_full_name(),
            'vehicles': vehicles_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving vehicles: {str(e)}'
        })


def assignment_report(request):
    """Generate assignment report"""
    unassigned_driver = get_or_create_unassigned_driver()

    vehicles = Vehicle.objects.select_related('driver__user').annotate(
        active_shipments_count=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
    ).order_by('plate_number')

    drivers = Driver.objects.select_related('user').filter(
        user__is_active=True
    ).annotate(
        vehicle_count=Count('vehicle', filter=~Q(vehicle__driver=unassigned_driver) if unassigned_driver else Q()),
        active_shipments_count=Count('vehicle__shipment',
                                     filter=Q(vehicle__shipment__status__in=['pending', 'in_transit']))
    ).order_by('user__first_name')

    context = {
        'vehicles': vehicles.exclude(driver=unassigned_driver) if unassigned_driver else vehicles,
        'drivers': drivers,
        'unassigned_vehicles': vehicles.filter(driver=unassigned_driver) if unassigned_driver else [],
        'report_date': timezone.now(),
        'total_vehicles': vehicles.count(),
        'assigned_vehicles': vehicles.exclude(
            driver=unassigned_driver).count() if unassigned_driver else vehicles.count(),
        'unassigned_vehicles_count': vehicles.filter(driver=unassigned_driver).count() if unassigned_driver else 0,
        'active_drivers': drivers.filter(vehicle_count__gt=0).count(),
        'drivers_without_vehicles': drivers.filter(vehicle_count=0).count(),
    }

    return render(request, 'logistics/assignment_report.html', context)


class AssignmentAnalyticsView(LoginRequiredMixin, ListView):
    """Advanced analytics view for assignment performance"""
    template_name = 'logistics/assignment_analytics.html'

    def get(self, request, *args, **kwargs):
        # Time-based analysis
        now = timezone.now()
        last_30_days = now - timedelta(days=30)
        last_7_days = now - timedelta(days=7)

        # Get unassigned driver to exclude from calculations
        unassigned_driver = get_or_create_unassigned_driver()

        # Assignment trends
        shipments_last_30_days = Shipment.objects.filter(
            created_at__gte=last_30_days
        ).count()

        shipments_last_7_days = Shipment.objects.filter(
            created_at__gte=last_7_days
        ).count()

        # Driver performance metrics
        driver_performance = Driver.objects.select_related('user').filter(
            user__is_active=True
        ).annotate(
            total_shipments=Count('shipment'),
            recent_shipments=Count('shipment', filter=Q(shipment__created_at__gte=last_30_days))
        ).order_by('-total_shipments')

        # Vehicle utilization over time
        vehicle_utilization = Vehicle.objects.select_related('driver__user').annotate(
            total_shipments=Count('shipment'),
            active_shipments=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit'])),
            utilization_score=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
        ).order_by('-utilization_score')

        # Filter out unassigned driver vehicles
        if unassigned_driver:
            vehicle_utilization = vehicle_utilization.exclude(driver=unassigned_driver)

        context = {
            'shipments_last_30_days': shipments_last_30_days,
            'shipments_last_7_days': shipments_last_7_days,
            'driver_performance': driver_performance[:10],  # Top 10
            'vehicle_utilization': vehicle_utilization[:10],  # Top 10
            'report_date': now,
        }

        return render(request, self.template_name, context)


@login_required
def assignment_report_export(request):
    """Export assignment report as CSV"""
    response = HttpResponse(content_type='text/csv')
    response[
        'Content-Disposition'] = f'attachment; filename="assignment_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    unassigned_driver = get_or_create_unassigned_driver()

    # Write header
    writer.writerow(['Driver-Vehicle Assignment Report'])
    writer.writerow([f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])

    # Vehicle assignments
    writer.writerow(['Vehicle Assignments'])
    writer.writerow(
        ['Plate Number', 'Model', 'Capacity (kg)', 'Assigned Driver', 'Driver License', 'Status', 'Active Shipments'])

    vehicles = Vehicle.objects.select_related('driver__user').annotate(
        active_shipments_count=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
    ).order_by('plate_number')

    for vehicle in vehicles:
        if vehicle.driver and vehicle.driver != unassigned_driver:
            driver_name = vehicle.driver.user.get_full_name()
            driver_license = vehicle.driver.license_number
        else:
            driver_name = 'Unassigned'
            driver_license = 'N/A'

        status = 'In Use' if vehicle.active_shipments_count > 0 else 'Available'

        writer.writerow([
            vehicle.plate_number,
            vehicle.model,
            vehicle.capacity_kg,
            driver_name,
            driver_license,
            status,
            vehicle.active_shipments_count
        ])

    writer.writerow([])

    # Driver assignments
    writer.writerow(['Driver Assignments'])
    writer.writerow(['Driver Name', 'Email', 'Phone', 'License', 'Assigned Vehicles', 'Active Shipments', 'Join Date'])

    drivers = Driver.objects.select_related('user').filter(
        user__is_active=True
    ).annotate(
        vehicle_count=Count('vehicle', filter=~Q(vehicle__driver=unassigned_driver) if unassigned_driver else Q()),
        active_shipments_count=Count('vehicle__shipment',
                                     filter=Q(vehicle__shipment__status__in=['pending', 'in_transit']))
    ).order_by('user__first_name')

    for driver in drivers:
        writer.writerow([
            driver.user.get_full_name(),
            driver.user.email,
            driver.phone,
            driver.license_number,
            driver.vehicle_count,
            driver.active_shipments_count,
            driver.user.date_joined.strftime('%Y-%m-%d')
        ])

    return response


@login_required
def assignment_report_json(request):
    """Return assignment report data as JSON for API consumption"""
    unassigned_driver = get_or_create_unassigned_driver()

    # Get vehicle data
    vehicles = Vehicle.objects.select_related('driver__user').annotate(
        active_shipments_count=Count('shipment', filter=Q(shipment__status__in=['pending', 'in_transit']))
    ).order_by('plate_number')

    # Get driver data
    drivers = Driver.objects.select_related('user').filter(
        user__is_active=True
    ).annotate(
        vehicle_count=Count('vehicle', filter=~Q(vehicle__driver=unassigned_driver) if unassigned_driver else Q()),
        active_shipments_count=Count('vehicle__shipment',
                                     filter=Q(vehicle__shipment__status__in=['pending', 'in_transit']))
    ).order_by('user__first_name')

    # Prepare vehicle data
    vehicles_data = []
    for vehicle in vehicles:
        if vehicle.driver and vehicle.driver != unassigned_driver:
            driver_info = {
                'id': vehicle.driver.id,
                'name': vehicle.driver.user.get_full_name(),
                'license': vehicle.driver.license_number,
                'phone': vehicle.driver.phone,
                'email': vehicle.driver.user.email
            }
        else:
            driver_info = None

        vehicles_data.append({
            'id': vehicle.id,
            'plate_number': vehicle.plate_number,
            'model': vehicle.model,
            'capacity_kg': vehicle.capacity_kg,
            'driver': driver_info,
            'status': 'in_use' if vehicle.active_shipments_count > 0 else 'available',
            'active_shipments': vehicle.active_shipments_count
        })

    # Prepare driver data
    drivers_data = []
    for driver in drivers:
        drivers_data.append({
            'id': driver.id,
            'name': driver.user.get_full_name(),
            'email': driver.user.email,
            'phone': driver.phone,
            'license': driver.license_number,
            'vehicle_count': driver.vehicle_count,
            'active_shipments': driver.active_shipments_count,
            'join_date': driver.user.date_joined.isoformat()
        })

    # Calculate metrics
    total_vehicles = len(vehicles_data)
    assigned_vehicles = len([v for v in vehicles_data if v['driver'] is not None])
    unassigned_vehicles = total_vehicles - assigned_vehicles

    data = {
        'report_date': timezone.now().isoformat(),
        'summary': {
            'total_vehicles': total_vehicles,
            'assigned_vehicles': assigned_vehicles,
            'unassigned_vehicles': unassigned_vehicles,
            'total_drivers': len(drivers_data),
            'active_drivers': len([d for d in drivers_data if d['vehicle_count'] > 0]),
            'assignment_efficiency': (assigned_vehicles / total_vehicles * 100) if total_vehicles > 0 else 0
        },
        'vehicles': vehicles_data,
        'drivers': drivers_data
    }

    return JsonResponse(data, indent=2)


def export_shipments_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="shipments.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Destination', 'Driver', 'Vehicle', 'Weight (kg)', 'Type', 'Status', 'Created'])

    for shipment in Shipment.objects.all():
        writer.writerow([
            shipment.id,
            f"{shipment.shipping_address.city}, {shipment.shipping_address.region}",
            shipment.driver.user.get_full_name() if shipment.driver else '',
            shipment.vehicle.plate_number if shipment.vehicle else '',
            shipment.weight_kg,
            shipment.get_shipment_type_display(),
            shipment.get_status_display(),
            shipment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    return response

def export_shipments_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipments"

    columns = ['ID', 'Destination', 'Driver', 'Vehicle', 'Weight (kg)', 'Type', 'Status', 'Created']
    ws.append(columns)

    for shipment in Shipment.objects.all():
        ws.append([
            shipment.id,
            f"{shipment.shipping_address.city}, {shipment.shipping_address.region}",
            shipment.driver.user.get_full_name() if shipment.driver else '',
            shipment.vehicle.plate_number if shipment.vehicle else '',
            shipment.weight_kg,
            shipment.get_shipment_type_display(),
            shipment.get_status_display(),
            shipment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="shipments.xlsx"'
    wb.save(response)
    return response
